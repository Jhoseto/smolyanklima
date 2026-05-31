#!/usr/bin/env python3
"""
Импорт на исторически продажби „втора употреба“ от Klimatici *.xlsx (2023, 2022, vtora).

2023 (Klimatici 2023.xlsx) — ред 3+, колони A–F:
  A: вътрешно, B: външно, C: клиент, D: дата, E: цена/бележка, F: доп. бележка

2022 (Klimatici vtora - 2022.xlsx) — ред 2+, колони B–G:
  B: външно (модел), C: вътрешно, E: клиент, F: дата, G: цена

vtora (Klimatici vtora.xlsx) — ред 2+, колони B–C + P–T:
  B: външно, C: вътрешно, P: забележки, Q: дата, R: клиент, S: цена, T: гаранция

Използване:
  python backend/scripts/import_klimatici2023_used_sales.py --year 2023 --preview
  python backend/scripts/import_klimatici2023_used_sales.py --year vtora --sql
  python backend/scripts/import_klimatici2023_used_sales.py --stock --preview
  python backend/scripts/import_klimatici2023_used_sales.py --stock --sql

Изход:
  --preview → Doc/Klimatici{year}_used_import_preview.tsv + warnings
  --sql     → backend/supabase/seeds/0009|0011_klimatici{year}_used_historical_sales.sql
  --stock   → Doc/Klimatici_used_stock_preview.tsv + seeds/0019_klimatici_used_stock_inventory.sql
              (редове без продажба → product_condition=used, stock_status=in_stock)
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DOC_DIR = ROOT.parent / "Doc" / "VTORA" / "VTORA" / "Old"
STOCK_PREVIEW_TSV = ROOT.parent / "Doc" / "Klimatici_used_stock_preview.tsv"
STOCK_WARNINGS_TXT = ROOT.parent / "Doc" / "Klimatici_used_stock_warnings.txt"
STOCK_OUT_SQL = ROOT / "backend" / "supabase" / "seeds" / "0019_klimatici_used_stock_inventory.sql"
STOCK_ROLLBACK_SEED = "0020_rollback_klimatici_used_stock.sql"

@dataclass(frozen=True)
class ImportConfig:
    year: str
    xlsx_path: Path
    preview_tsv: Path
    warnings_txt: Path
    out_sql: Path
    rollback_seed: str
    import_label: str
    slug_prefix: str
    stock_slug_prefix: str


IMPORT_CONFIGS: dict[str, ImportConfig] = {
    "2023": ImportConfig(
        year="2023",
        xlsx_path=DOC_DIR / "Klimatici 2023.xlsx",
        preview_tsv=ROOT.parent / "Doc" / "Klimatici2023_used_import_preview.tsv",
        warnings_txt=ROOT.parent / "Doc" / "Klimatici2023_used_import_warnings.txt",
        out_sql=ROOT / "backend" / "supabase" / "seeds" / "0009_klimatici2023_used_historical_sales.sql",
        rollback_seed="0010_rollback_klimatici2023_used_sales.sql",
        import_label="Klimatici2023 VTORA",
        slug_prefix="klimatici2023-used",
        stock_slug_prefix="klimatici2023-stock",
    ),
    "2022": ImportConfig(
        year="2022",
        xlsx_path=DOC_DIR / "Klimatici vtora - 2022.xlsx",
        preview_tsv=ROOT.parent / "Doc" / "Klimatici2022_used_import_preview.tsv",
        warnings_txt=ROOT.parent / "Doc" / "Klimatici2022_used_import_warnings.txt",
        out_sql=ROOT / "backend" / "supabase" / "seeds" / "0011_klimatici2022_used_historical_sales.sql",
        rollback_seed="0012_rollback_klimatici2022_used_sales.sql",
        import_label="Klimatici2022 VTORA",
        slug_prefix="klimatici2022-used",
        stock_slug_prefix="klimatici2022-stock",
    ),
    "vtora": ImportConfig(
        year="vtora",
        xlsx_path=DOC_DIR / "Klimatici vtora.xlsx",
        preview_tsv=ROOT.parent / "Doc" / "KlimaticiVtora_used_import_preview.tsv",
        warnings_txt=ROOT.parent / "Doc" / "KlimaticiVtora_used_import_warnings.txt",
        out_sql=ROOT / "backend" / "supabase" / "seeds" / "0014_klimatici_vtora_used_historical_sales.sql",
        rollback_seed="0015_rollback_klimatici_vtora_used_sales.sql",
        import_label="KlimaticiVtora VTORA",
        slug_prefix="klimatici-vtora-used",
        stock_slug_prefix="klimatici-vtora-stock",
    ),
}

SHEET_BRAND: dict[str, str] = {
    "toshiba": "Toshiba",
    "mitsubishi": "Mitsubishi Electric",
    "nacional": "Nacional",
    "hitachi": "Hitachi",
    "panasonic": "Panasonic",
    "sharp": "Sharp",
    "sanyo": "Sanyo",
    "fujitsu": "Fujitsu",
    "daikin": "Daikin",
}

PHONE_PATTERNS = [
    re.compile(r"(\+?\d{12,15})"),
    re.compile(r"(\b0\d{9}\b)"),
]

PLACES = {
    "СОФИЯ", "ПЛОВДИВ", "СМОЛЯН", "РУДОЗЕМ", "ДЕВИН", "ПАМПОРОВО",
    "БАНИТЕ", "МАДАН", "ЧЕПЕЛАРЕ", "БЯЛА РЕКА", "ДАВИДКОВО",
    "МОМЧИЛОВЦИ", "НАСТАН", "ЛЕСКА", "МОГИЛИЦА", "ПАВЕЛСКО",
    "ВИШНЕВО", "ГЪЛЪБОВО", "ЛЯСКОВО", "ГЪРЦИЯ", "ЛЕНОВО",
    "ВЪРБИНА", "БОСТИНА", "ЧАЛА", "УСТОВО", "ПЕРСЕНК",
    "КОС СМОЛЯН", "БЯЛА", "ФАТОВО", "ВАСИЛ ЛЕВСКИ", "КОСИТЕ",
    "ЛЪКИ", "НЕДЕЛИНО", "ЕЛХОВЕЦ", "КАТРАНИЦА", "КЪРДЖАЛИ",
    "РАВНИЩА", "ЛОВЦИ", "ВАРНА", "RUDOZEM",
}

SKIP_CLIENT_ONLY = {
    "ПЛОВДИВ",
    "СОФИЯ",
    "СМОЛЯН",
}


@dataclass
class ParsedRow:
    sheet_name: str
    sheet_row: int
    brand_db: str
    model: str
    indoor_serial: str
    outdoor_serial: str
    sale_date: str | None
    sale_price: float | None
    client_raw: str
    client_name: str
    client_phone: str | None
    client_address: str
    note_col5: str
    note_col6: str
    warnings: list[str] = field(default_factory=list)


@dataclass
class ParsedStockRow:
    source_year: str
    import_label: str
    stock_slug_prefix: str
    sheet_name: str
    sheet_row: int
    brand_db: str
    model: str
    indoor_serial: str
    outdoor_serial: str
    purchase_price: float | None
    list_price: float | None
    extra_note: str
    warnings: list[str] = field(default_factory=list)


JUNK_SERIALS = {"няма табелка", "брак", "brak"}


def cell_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def normalize_phone(raw: str) -> str:
    p = re.sub(r"[^\d+]", "", raw)
    if not p:
        return ""
    if p.startswith("00"):
        p = "+" + p[2:]
    if p.startswith("+"):
        if p.startswith("+359") and len(p) == 13:
            return "0" + p[4:]
        return p
    if p.startswith("359") and len(p) == 12:
        return "0" + p[3:]
    if p.startswith("0") and len(p) == 10:
        return p
    if len(p) == 9:
        return "0" + p
    return p


def format_phone_display(canonical: str) -> str:
    if canonical.startswith("+"):
        return canonical
    digits = re.sub(r"[^\d]", "", canonical)
    if digits.startswith("359") and len(digits) == 12:
        digits = "0" + digits[3:]
    if digits.startswith("0") and len(digits) == 10:
        return f"{digits[:4]} {digits[4:7]} {digits[7:]}"
    return canonical


def extract_phones(text: str) -> list[str]:
    found: list[str] = []
    for pat in PHONE_PATTERNS:
        for m in pat.finditer(text):
            found.append(m.group(1))
    seen: set[str] = set()
    uniq: list[str] = []
    for p in found:
        if p not in seen:
            seen.add(p)
            uniq.append(p)
    return uniq


def clean_text(text: str) -> str:
    t = re.sub(r"\s+", " ", (text or "").replace("\n", " ")).strip()
    return t


def split_name_address(text: str) -> tuple[str, str]:
    parts = [p.strip() for p in re.split(r"[,;]", text) if p.strip()]
    if not parts:
        return "", ""
    if len(parts) == 1:
        upper = parts[0].upper()
        if upper in PLACES:
            return "", parts[0]
        return parts[0], ""
    last = parts[-1]
    if any(pl in last.upper() for pl in PLACES):
        return ", ".join(parts[:-1]), last
    return ", ".join(parts), ""


def parse_customer_text(raw: str) -> tuple[str, str | None, str]:
    text = clean_text(raw)
    if not text:
        return "", None, ""
    phones = extract_phones(text)
    residue = text
    for p in phones:
        residue = residue.replace(p, "")
    residue = clean_text(residue)
    name, address = split_name_address(residue)
    phone = normalize_phone(phones[0]) if phones else None
    display_phone = format_phone_display(phone) if phone else None
    if not name and address:
        name = address
        address = ""
    return name.strip(), display_phone, address.strip()


def parse_price(val) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        n = round(float(val), 2)
        return n if n > 0 else None
    s = str(val).strip().upper()
    s = s.replace("ЕВРО", "").replace("EUR", "").replace("€", "").replace("?", "")
    m = re.search(r"\d+(?:[.,]\d+)?", s)
    if m:
        return round(float(m.group().replace(",", ".")), 2)
    return None


def parse_date(val) -> str | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    if not s:
        return None
    m = re.match(r"(\d{1,2})[,\./](\d{1,2})[,\./]'?(\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000 if y < 50 else 1900
        try:
            return datetime(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def is_price_like(val) -> bool:
    if val is None or val == "":
        return False
    if isinstance(val, (int, float)):
        return True
    s = str(val).strip()
    if re.fullmatch(r"\d+(?:[.,]\d+)?\??", s):
        return True
    if re.search(r"\d", s) and re.search(r"(^|\s)\d{3,4}(\s*[Ee€])?\s*$", s):
        return True
    return parse_price(val) is not None


def note_text(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return ""
    s = clean_text(str(val))
    if is_price_like(val):
        return ""
    return s


def brand_from_sheet(sheet_name: str) -> str | None:
    key = re.sub(r"\s*вън\s*$", "", sheet_name, flags=re.IGNORECASE).strip().lower()
    return SHEET_BRAND.get(key)


def should_import(client_raw: str, sale_date: str | None, price_val, extra_note=None, inventory_note=None) -> bool:
    client = clean_text(client_raw)
    upper = client.upper()
    if upper in SKIP_CLIENT_ONLY:
        return False
    if "BRAK" in upper or "БРАК" in upper:
        return False
    if "ВЪТРЕШНА ПЛАТКА" in upper and not extract_phones(client):
        return False
    if inventory_note:
        note_upper = clean_text(str(inventory_note)).upper()
        if "BRAK" in note_upper or "БРАК" in note_upper:
            return False
        if "ВЪТРЕШНА ПЛАТКА" in note_upper and not extract_phones(client):
            return False
    if "ЗА ЧАСТИ" in upper and not extract_phones(client):
        return False
    price = parse_price(price_val)
    has_client = bool(client) and upper not in SKIP_CLIENT_ONLY
    has_phone = bool(extract_phones(client))
    has_date = sale_date is not None
    has_price = price is not None
    if has_client and (has_phone or has_date or has_price):
        return True
    if has_date and has_price:
        return True
    if extra_note and parse_price(extra_note) is None and clean_text(str(extra_note)):
        return has_client
    return False


def is_plausible_serial(value: str) -> bool:
    serial = clean_text(value)
    if len(serial) < 3:
        return False
    if serial.lower() in JUNK_SERIALS:
        return False
    return bool(re.search(r"[A-Za-z]", serial))


def has_plausible_stock_serials(indoor: str, outdoor: str) -> bool:
    return is_plausible_serial(indoor) or is_plausible_serial(outdoor)


def should_include_as_stock(
    *,
    indoor: str,
    outdoor: str,
    client_raw: str,
    sale_date: str | None,
    price_val,
    extra_note=None,
    inventory_note=None,
) -> bool:
    if not has_plausible_stock_serials(indoor, outdoor):
        return False
    if should_import(client_raw, sale_date, price_val, extra_note, inventory_note):
        return False
    client = clean_text(client_raw)
    upper = client.upper()
    if "BRAK" in upper or "БРАК" in upper:
        return False
    for note in (inventory_note, extra_note):
        if note:
            note_upper = clean_text(str(note)).upper()
            if "BRAK" in note_upper or "БРАК" in note_upper:
                return False
    return True


def append_stock_row(
    rows: list[ParsedStockRow],
    stats: Counter,
    config: ImportConfig,
    *,
    sheet_name: str,
    sheet_row: int,
    brand_db: str,
    indoor: str,
    outdoor: str,
    purchase_price: float | None,
    list_price: float | None,
    extra_note: str = "",
) -> None:
    model = indoor or outdoor
    warnings: list[str] = []
    if purchase_price is None and list_price is None:
        warnings.append("missing_prices")
    rows.append(
        ParsedStockRow(
            source_year=config.year,
            import_label=config.import_label,
            stock_slug_prefix=config.stock_slug_prefix,
            sheet_name=sheet_name,
            sheet_row=sheet_row,
            brand_db=brand_db,
            model=model,
            indoor_serial=indoor,
            outdoor_serial=outdoor,
            purchase_price=purchase_price,
            list_price=list_price,
            extra_note=extra_note,
            warnings=warnings,
        )
    )
    stats[f"stock:{config.year}:{sheet_name}"] += 1


def append_row(
    rows: list[ParsedRow],
    stats: Counter,
    *,
    sheet_name: str,
    sheet_row: int,
    brand_db: str,
    indoor: str,
    outdoor: str,
    client_raw: str,
    sale_date: str | None,
    price_val,
    note_val=None,
    extra_note_val=None,
) -> None:
    warnings: list[str] = []
    sale_price = parse_price(price_val)
    if sale_price is None and client_raw and re.search(r"(лв|евро|eur|€)", client_raw, re.I):
        sale_price = parse_price(client_raw)
    if sale_price is not None and sale_price >= 10000:
        corrected = round(sale_price / 10, 2)
        warnings.append(f"price_typo_corrected:{sale_price}->{corrected}")
        sale_price = corrected
    n5 = note_text(note_val if note_val is not None else price_val)
    n6 = note_text(extra_note_val)
    client_name, client_phone, client_address = parse_customer_text(client_raw)

    if sale_price is None:
        warnings.append("missing_sale_price")
    if sale_date is None:
        warnings.append("missing_sale_date")
    if not client_name and not client_phone:
        warnings.append("missing_client")
    if sale_price is not None and sale_price > 10000:
        warnings.append(f"suspicious_price:{sale_price}")

    model = indoor or outdoor
    rows.append(
        ParsedRow(
            sheet_name=sheet_name,
            sheet_row=sheet_row,
            brand_db=brand_db,
            model=model,
            indoor_serial=indoor,
            outdoor_serial=outdoor,
            sale_date=sale_date,
            sale_price=sale_price,
            client_raw=client_raw,
            client_name=client_name or client_raw or "—",
            client_phone=client_phone,
            client_address=client_address,
            note_col5=n5,
            note_col6=n6,
            warnings=warnings,
        )
    )
    stats[f"sheet:{sheet_name}"] += 1


def parse_workbook_2023(wb, rows: list[ParsedRow], stats: Counter) -> None:
    for sheet_name in wb.sheetnames:
        brand_db = brand_from_sheet(sheet_name)
        if not brand_db:
            stats[f"unknown_sheet:{sheet_name}"] += 1
            continue
        ws = wb[sheet_name]
        for r in range(3, ws.max_row + 1):
            indoor = cell_str(ws.cell(r, 1).value)
            outdoor = cell_str(ws.cell(r, 2).value)
            client_raw = cell_str(ws.cell(r, 3).value)
            sale_date = parse_date(ws.cell(r, 4).value)
            col5 = ws.cell(r, 5).value
            col6 = ws.cell(r, 6).value if ws.max_column >= 6 else None

            if not indoor and not outdoor:
                continue
            if not should_import(client_raw, sale_date, col5, col6):
                stats["skipped_inventory"] += 1
                continue

            append_row(
                rows, stats,
                sheet_name=sheet_name, sheet_row=r, brand_db=brand_db,
                indoor=indoor, outdoor=outdoor, client_raw=client_raw,
                sale_date=sale_date, price_val=col5, note_val=col5, extra_note_val=col6,
            )


def parse_workbook_2022(wb, rows: list[ParsedRow], stats: Counter) -> None:
    for sheet_name in wb.sheetnames:
        brand_db = brand_from_sheet(sheet_name)
        if not brand_db:
            stats[f"unknown_sheet:{sheet_name}"] += 1
            continue
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            outdoor = cell_str(ws.cell(r, 2).value)
            indoor = cell_str(ws.cell(r, 3).value)
            client_raw = cell_str(ws.cell(r, 5).value)
            sale_date = parse_date(ws.cell(r, 6).value)
            price_val = ws.cell(r, 7).value

            if not indoor and not outdoor:
                continue
            if not should_import(client_raw, sale_date, price_val):
                stats["skipped_inventory"] += 1
                continue

            append_row(
                rows, stats,
                sheet_name=sheet_name, sheet_row=r, brand_db=brand_db,
                indoor=indoor, outdoor=outdoor, client_raw=client_raw,
                sale_date=sale_date, price_val=price_val,
            )


def parse_workbook_vtora(wb, rows: list[ParsedRow], stats: Counter) -> None:
    for sheet_name in wb.sheetnames:
        brand_db = brand_from_sheet(sheet_name)
        if not brand_db:
            stats[f"unknown_sheet:{sheet_name}"] += 1
            continue
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            outdoor = cell_str(ws.cell(r, 2).value)
            indoor = cell_str(ws.cell(r, 3).value)
            inventory_note = ws.cell(r, 16).value
            sale_date = parse_date(ws.cell(r, 17).value)
            client_raw = cell_str(ws.cell(r, 18).value)
            price_val = ws.cell(r, 19).value
            warranty_note = ws.cell(r, 20).value if ws.max_column >= 20 else None

            if not indoor and not outdoor:
                continue
            if not should_import(client_raw, sale_date, price_val, inventory_note=inventory_note):
                stats["skipped_inventory"] += 1
                continue

            extra_notes = [inventory_note, warranty_note]
            extra_note_val = " · ".join(
                note_text(n) for n in extra_notes if note_text(n)
            ) or None

            append_row(
                rows, stats,
                sheet_name=sheet_name, sheet_row=r, brand_db=brand_db,
                indoor=indoor, outdoor=outdoor, client_raw=client_raw,
                sale_date=sale_date, price_val=price_val,
                extra_note_val=extra_note_val,
            )


def parse_workbook(config: ImportConfig) -> tuple[list[ParsedRow], Counter]:
    wb = openpyxl.load_workbook(str(config.xlsx_path), data_only=True)
    rows: list[ParsedRow] = []
    stats: Counter = Counter()

    if config.year == "2022":
        parse_workbook_2022(wb, rows, stats)
    elif config.year == "vtora":
        parse_workbook_vtora(wb, rows, stats)
    else:
        parse_workbook_2023(wb, rows, stats)

    return rows, stats


def parse_stock_workbook_2023(wb, rows: list[ParsedStockRow], stats: Counter, config: ImportConfig) -> None:
    for sheet_name in wb.sheetnames:
        brand_db = brand_from_sheet(sheet_name)
        if not brand_db:
            stats[f"unknown_sheet:{sheet_name}"] += 1
            continue
        ws = wb[sheet_name]
        for r in range(3, ws.max_row + 1):
            indoor = cell_str(ws.cell(r, 1).value)
            outdoor = cell_str(ws.cell(r, 2).value)
            client_raw = cell_str(ws.cell(r, 3).value)
            sale_date = parse_date(ws.cell(r, 4).value)
            col5 = ws.cell(r, 5).value
            col6 = ws.cell(r, 6).value if ws.max_column >= 6 else None
            if not should_include_as_stock(
                indoor=indoor,
                outdoor=outdoor,
                client_raw=client_raw,
                sale_date=sale_date,
                price_val=col5,
                extra_note=col6,
            ):
                continue
            extra = " · ".join(
                note for note in (note_text(col5), note_text(col6), clean_text(client_raw)) if note
            )
            append_stock_row(
                rows, stats, config,
                sheet_name=sheet_name, sheet_row=r, brand_db=brand_db,
                indoor=indoor, outdoor=outdoor,
                purchase_price=None, list_price=None, extra_note=extra,
            )


def parse_stock_workbook_2022(wb, rows: list[ParsedStockRow], stats: Counter, config: ImportConfig) -> None:
    for sheet_name in wb.sheetnames:
        brand_db = brand_from_sheet(sheet_name)
        if not brand_db:
            stats[f"unknown_sheet:{sheet_name}"] += 1
            continue
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            outdoor = cell_str(ws.cell(r, 2).value)
            indoor = cell_str(ws.cell(r, 3).value)
            purchase_price = parse_price(ws.cell(r, 4).value)
            client_raw = cell_str(ws.cell(r, 5).value)
            sale_date = parse_date(ws.cell(r, 6).value)
            price_val = ws.cell(r, 7).value
            if not should_include_as_stock(
                indoor=indoor,
                outdoor=outdoor,
                client_raw=client_raw,
                sale_date=sale_date,
                price_val=price_val,
            ):
                continue
            list_price = parse_price(price_val)
            extra = clean_text(client_raw)
            append_stock_row(
                rows, stats, config,
                sheet_name=sheet_name, sheet_row=r, brand_db=brand_db,
                indoor=indoor, outdoor=outdoor,
                purchase_price=purchase_price, list_price=list_price, extra_note=extra,
            )


def parse_stock_workbook_vtora(wb, rows: list[ParsedStockRow], stats: Counter, config: ImportConfig) -> None:
    for sheet_name in wb.sheetnames:
        brand_db = brand_from_sheet(sheet_name)
        if not brand_db:
            stats[f"unknown_sheet:{sheet_name}"] += 1
            continue
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            outdoor = cell_str(ws.cell(r, 2).value)
            indoor = cell_str(ws.cell(r, 3).value)
            purchase_price = parse_price(ws.cell(r, 4).value)
            list_price = parse_price(ws.cell(r, 5).value)
            inventory_note = ws.cell(r, 16).value
            sale_date = parse_date(ws.cell(r, 17).value)
            client_raw = cell_str(ws.cell(r, 18).value)
            price_val = ws.cell(r, 19).value
            warranty_note = ws.cell(r, 20).value if ws.max_column >= 20 else None
            if not should_include_as_stock(
                indoor=indoor,
                outdoor=outdoor,
                client_raw=client_raw,
                sale_date=sale_date,
                price_val=price_val,
                inventory_note=inventory_note,
            ):
                continue
            extra_notes = [inventory_note, warranty_note, client_raw]
            extra = " · ".join(note_text(n) for n in extra_notes if note_text(n)) or ""
            append_stock_row(
                rows, stats, config,
                sheet_name=sheet_name, sheet_row=r, brand_db=brand_db,
                indoor=indoor, outdoor=outdoor,
                purchase_price=purchase_price, list_price=list_price, extra_note=extra,
            )


def parse_stock_workbook(config: ImportConfig) -> tuple[list[ParsedStockRow], Counter]:
    wb = openpyxl.load_workbook(str(config.xlsx_path), data_only=True)
    rows: list[ParsedStockRow] = []
    stats: Counter = Counter()

    if config.year == "2022":
        parse_stock_workbook_2022(wb, rows, stats, config)
    elif config.year == "vtora":
        parse_stock_workbook_vtora(wb, rows, stats, config)
    else:
        parse_stock_workbook_2023(wb, rows, stats, config)

    return rows, stats


def write_preview(rows: list[ParsedRow], stats: Counter, config: ImportConfig) -> None:
    config.preview_tsv.parent.mkdir(parents=True, exist_ok=True)
    with config.preview_tsv.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow([
            "sheet", "row", "brand", "model", "indoor", "outdoor",
            "sale_date", "sale_price", "client_name", "client_phone", "client_address",
            "note5", "note6", "warnings",
        ])
        for row in rows:
            w.writerow([
                row.sheet_name, row.sheet_row, row.brand_db, row.model,
                row.indoor_serial, row.outdoor_serial,
                row.sale_date or "", row.sale_price if row.sale_price is not None else "",
                row.client_name, row.client_phone or "", row.client_address,
                row.note_col5, row.note_col6, ";".join(row.warnings),
            ])

    warn_counter = Counter()
    for row in rows:
        for w in row.warnings:
            warn_counter[w.split(":", 1)[0]] += 1

    lines = [
        f"Source: {config.xlsx_path}",
        f"Importable rows: {len(rows)}",
        "",
        "Per sheet:",
        *[f"  {k}: {v}" for k, v in sorted(stats.items())],
        "",
        "Warnings:",
        *[f"  {k}: {v}" for k, v in warn_counter.most_common()],
    ]
    config.warnings_txt.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Preview TSV: {config.preview_tsv}")
    print(f"Warnings:    {config.warnings_txt}")
    print(f"Rows: {len(rows)}")


def sql_escape(value) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    s = str(value).replace("\\", "\\\\").replace("'", "''")
    return f"'{s}'"


def sql_lit(value, pg_type: str) -> str:
    if value is None:
        return f"NULL::{pg_type}"
    if pg_type == "int":
        return str(int(value))
    if pg_type == "numeric(10,2)":
        return f"{float(value)}::numeric(10,2)"
    if pg_type == "date":
        return f"{sql_escape(value)}::date"
    return sql_escape(value)


def build_notes(row: ParsedRow, import_label: str) -> str:
    parts = [f"Импорт {import_label}, лист {row.sheet_name} ред {row.sheet_row}"]
    for note in (row.note_col5, row.note_col6):
        if note:
            parts.append(note)
    return " · ".join(parts)


def write_sql(rows: list[ParsedRow], config: ImportConfig) -> None:
    config.out_sql.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    lines.append("-- =====================================================================")
    lines.append(f"-- Seed: Исторически продажби „втора употреба“ от Klimatici {config.year}.xlsx")
    lines.append("-- =====================================================================")
    lines.append(f"-- Редове: {len(rows)}")
    lines.append("-- Продукти: product_condition = used")
    lines.append(f"-- Идемпотентност: notes LIKE 'Импорт {config.import_label}, лист % ред N'")
    lines.append(f"-- Rollback: seeds/{config.rollback_seed}")
    lines.append("-- ВАЖНО: Един DO блок — пусни целия файл (Ctrl+A → Run).")
    lines.append("-- =====================================================================")
    lines.append("")
    lines.append("DO $import$")
    lines.append("DECLARE")
    lines.append("  r RECORD;")
    lines.append("  v_brand_id uuid;")
    lines.append("  v_type_id uuid;")
    lines.append("  v_contact_id uuid;")
    lines.append("  v_product_id uuid;")
    lines.append("  v_sale_id uuid;")
    lines.append("  v_slug text;")
    lines.append("  v_name text;")
    lines.append("  v_note text;")
    lines.append("  v_imported int := 0;")
    lines.append("  v_skipped int := 0;")
    lines.append("BEGIN")
    lines.append("  INSERT INTO public.brands (slug, name, color, is_active)")
    lines.append("  VALUES ('sanyo', 'Sanyo', '#1D4ED8', true)")
    lines.append("  ON CONFLICT (slug) DO UPDATE SET is_active = excluded.is_active;")
    lines.append("")
    lines.append("  SELECT id INTO v_type_id FROM public.product_types ORDER BY name LIMIT 1;")
    lines.append("  IF v_type_id IS NULL THEN")
    lines.append("    RAISE EXCEPTION 'Липсва product_types seed.';")
    lines.append("  END IF;")
    lines.append("")
    lines.append("  FOR r IN")
    lines.append("    SELECT *")
    lines.append("    FROM (")
    lines.append("      VALUES")

    value_lines: list[str] = []
    for row in rows:
        value_lines.append(
            "        ("
            + ", ".join([
                sql_lit(row.sheet_name, "text"),
                sql_lit(row.sheet_row, "int"),
                sql_lit(row.brand_db, "text"),
                sql_lit(row.model, "text"),
                sql_lit(row.indoor_serial or None, "text"),
                sql_lit(row.outdoor_serial or None, "text"),
                sql_lit(row.sale_date, "date"),
                sql_lit(row.sale_price if row.sale_price is not None else 0, "numeric(10,2)"),
                sql_lit(row.client_name, "text"),
                sql_lit(row.client_phone, "text"),
                sql_lit(row.client_address or None, "text"),
                sql_lit(build_notes(row, config.import_label), "text"),
            ])
            + ")"
        )

    lines.append(",\n".join(value_lines))
    lines.append("    ) AS stage(")
    lines.append("      sheet_name, sheet_row, brand_name, model,")
    lines.append("      indoor_serial, outdoor_serial, sale_date, sale_price,")
    lines.append("      client_name, client_phone, client_address, notes")
    lines.append("    )")
    lines.append("    ORDER BY sheet_name, sheet_row")
    lines.append("  LOOP")
    lines.append("    v_product_id := NULL;")
    lines.append("    v_contact_id := NULL;")
    lines.append(f"    v_slug := '{config.slug_prefix}-' || r.sheet_name || '-' || r.sheet_row;")
    lines.append("    v_note := r.notes;")
    lines.append("")
    lines.append("    IF EXISTS (")
    lines.append("      SELECT 1 FROM public.work_items")
    lines.append("      WHERE event_code = 'sale'")
    lines.append(f"        AND notes LIKE 'Импорт {config.import_label}, лист ' || r.sheet_name || ' ред ' || r.sheet_row || '%'")
    lines.append("    ) THEN")
    lines.append("      v_skipped := v_skipped + 1;")
    lines.append("      CONTINUE;")
    lines.append("    END IF;")
    lines.append("    IF EXISTS (SELECT 1 FROM public.products WHERE slug = v_slug) THEN")
    lines.append("      v_skipped := v_skipped + 1;")
    lines.append("      CONTINUE;")
    lines.append("    END IF;")
    lines.append("")
    lines.append("    IF r.indoor_serial IS NOT NULL AND btrim(r.indoor_serial) <> '' THEN")
    lines.append("      SELECT p.id INTO v_product_id FROM public.products p")
    lines.append("      JOIN public.work_items w ON w.product_id = p.id AND w.event_code = 'sale'")
    lines.append("      WHERE upper(btrim(p.indoor_unit_serial)) = upper(btrim(r.indoor_serial))")
    lines.append("      LIMIT 1;")
    lines.append("    END IF;")
    lines.append("    IF v_product_id IS NULL AND r.outdoor_serial IS NOT NULL AND btrim(r.outdoor_serial) <> '' THEN")
    lines.append("      SELECT p.id INTO v_product_id FROM public.products p")
    lines.append("      JOIN public.work_items w ON w.product_id = p.id AND w.event_code = 'sale'")
    lines.append("      WHERE upper(btrim(p.outdoor_unit_serial)) = upper(btrim(r.outdoor_serial))")
    lines.append("      LIMIT 1;")
    lines.append("    END IF;")
    lines.append("    IF v_product_id IS NOT NULL THEN")
    lines.append("      v_skipped := v_skipped + 1;")
    lines.append("      CONTINUE;")
    lines.append("    END IF;")
    lines.append("")
    lines.append("    SELECT id INTO v_brand_id FROM public.brands WHERE name = r.brand_name LIMIT 1;")
    lines.append("    IF v_brand_id IS NULL THEN")
    lines.append(f"      RAISE WARNING 'Klimatici{config.year} % row %: липсва марка %', r.sheet_name, r.sheet_row, r.brand_name;")
    lines.append("      CONTINUE;")
    lines.append("    END IF;")
    lines.append("")
    lines.append("    IF r.client_phone IS NOT NULL AND length(btrim(r.client_phone)) >= 3 THEN")
    lines.append("      SELECT id INTO v_contact_id FROM public.contacts")
    lines.append("      WHERE phone = r.client_phone AND contact_kind = 'client' LIMIT 1;")
    lines.append("    END IF;")
    lines.append("    IF v_contact_id IS NULL THEN")
    lines.append("      SELECT id INTO v_contact_id FROM public.contacts")
    lines.append("      WHERE upper(btrim(full_name)) = upper(btrim(r.client_name))")
    lines.append("        AND contact_kind = 'client'")
    lines.append("        AND (r.client_phone IS NULL OR phone IS NULL OR phone = r.client_phone)")
    lines.append("      LIMIT 1;")
    lines.append("    END IF;")
    lines.append("    IF v_contact_id IS NULL THEN")
    lines.append("      INSERT INTO public.contacts (full_name, phone, address, contact_kind, customer_status)")
    lines.append("      VALUES (r.client_name, r.client_phone, r.client_address, 'client', 'active')")
    lines.append("      RETURNING id INTO v_contact_id;")
    lines.append("      IF r.client_phone IS NOT NULL AND length(btrim(r.client_phone)) >= 3 THEN")
    lines.append("        INSERT INTO public.contact_phones (contact_id, phone, is_primary, sort_order)")
    lines.append("        VALUES (v_contact_id, r.client_phone, true, 0) ON CONFLICT DO NOTHING;")
    lines.append("      END IF;")
    lines.append("    END IF;")
    lines.append("")
    lines.append("    v_name := r.brand_name || ' ' || coalesce(nullif(btrim(r.model), ''), 'климатик');")
    lines.append("")
    lines.append("    INSERT INTO public.products (")
    lines.append("      slug, name, brand_id, type_id, model_code, price, purchase_price,")
    lines.append("      indoor_unit_serial, outdoor_unit_serial, purchased_at,")
    lines.append("      product_condition, stock_status, stock_quantity, sold_quantity,")
    lines.append("      is_active, show_in_public_catalog")
    lines.append("    ) VALUES (")
    lines.append("      v_slug, v_name, v_brand_id, v_type_id, nullif(btrim(r.model), ''),")
    lines.append("      coalesce(r.sale_price, 0), NULL,")
    lines.append("      nullif(btrim(r.indoor_serial), ''), nullif(btrim(r.outdoor_serial), ''), r.sale_date,")
    lines.append("      'used', 'out_of_stock', 0, 1, false, false")
    lines.append("    ) RETURNING id INTO v_product_id;")
    lines.append("")
    lines.append("    INSERT INTO public.work_items (")
    lines.append("      type, event_code, status, priority, title, notes, due_date, completed_at,")
    lines.append("      product_id, contact_id, customer_name, customer_phone, customer_address,")
    lines.append("      quantity, unit_price, total_amount, purchase_price, sale_install_state")
    lines.append("    ) VALUES (")
    lines.append("      'sale', 'sale', 'done', 'medium',")
    lines.append("      'Продажба: ' || v_name,")
    lines.append("      v_note,")
    lines.append("      r.sale_date,")
    lines.append("      (coalesce(r.sale_date, current_date) + time '12:00:00') AT TIME ZONE 'Europe/Sofia',")
    lines.append("      v_product_id, v_contact_id, r.client_name, r.client_phone, r.client_address,")
    lines.append("      1, coalesce(r.sale_price, 0), coalesce(r.sale_price, 0), NULL, 'completed'")
    lines.append("    ) RETURNING id INTO v_sale_id;")
    lines.append("    v_imported := v_imported + 1;")
    lines.append("  END LOOP;")
    lines.append("")
    lines.append(f"  RAISE NOTICE 'Klimatici{config.year} used import: imported=%, skipped(existing)=%', v_imported, v_skipped;")
    lines.append("END")
    lines.append("$import$;")
    lines.append("")

    config.out_sql.write_text("\n".join(lines), encoding="utf-8")
    print(f"SQL seed: {config.out_sql}")
    print(f"Rows: {len(rows)}")


def build_stock_description(row: ParsedStockRow) -> str:
    parts = [f"Импорт {row.import_label} склад, лист {row.sheet_name} ред {row.sheet_row}"]
    if row.extra_note:
        parts.append(row.extra_note)
    return " · ".join(parts)


def write_stock_preview(rows: list[ParsedStockRow], stats: Counter) -> None:
    STOCK_PREVIEW_TSV.parent.mkdir(parents=True, exist_ok=True)
    with STOCK_PREVIEW_TSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow([
            "source", "sheet", "row", "brand", "model", "indoor", "outdoor",
            "purchase_price", "list_price", "extra_note", "warnings",
        ])
        for row in rows:
            w.writerow([
                row.source_year, row.sheet_name, row.sheet_row, row.brand_db, row.model,
                row.indoor_serial, row.outdoor_serial,
                row.purchase_price if row.purchase_price is not None else "",
                row.list_price if row.list_price is not None else "",
                row.extra_note, ";".join(row.warnings),
            ])

    warn_counter = Counter()
    for row in rows:
        for warning in row.warnings:
            warn_counter[warning.split(":", 1)[0]] += 1

    lines = [
        "Klimatici used stock import (all workbooks)",
        f"Importable rows: {len(rows)}",
        "",
        "Per source/sheet:",
        *[f"  {k}: {v}" for k, v in sorted(stats.items())],
        "",
        "Warnings:",
        *[f"  {k}: {v}" for k, v in sorted(warn_counter.items())],
    ]
    STOCK_WARNINGS_TXT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Preview TSV: {STOCK_PREVIEW_TSV}")
    print(f"Warnings: {STOCK_WARNINGS_TXT}")
    print(f"Stock rows: {len(rows)}")


def write_stock_sql(rows: list[ParsedStockRow]) -> None:
    STOCK_OUT_SQL.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    lines.append("-- =====================================================================")
    lines.append("-- Seed: Klimatici втора употреба — наличност (склад) от 2023/2022/vtora .xlsx")
    lines.append("-- =====================================================================")
    lines.append(f"-- Редове: {len(rows)}")
    lines.append("-- Продукти: product_condition = used, stock_status = in_stock")
    lines.append("-- Идемпотентност: slug + description LIKE 'Импорт % VTORA склад, лист % ред N'")
    lines.append(f"-- Rollback: seeds/{STOCK_ROLLBACK_SEED}")
    lines.append("-- ВАЖНО: Един DO блок — пусни целия файл (Ctrl+A → Run).")
    lines.append("-- =====================================================================")
    lines.append("")
    lines.append("DO $import$")
    lines.append("DECLARE")
    lines.append("  r RECORD;")
    lines.append("  v_brand_id uuid;")
    lines.append("  v_type_id uuid;")
    lines.append("  v_product_id uuid;")
    lines.append("  v_slug text;")
    lines.append("  v_name text;")
    lines.append("  v_desc text;")
    lines.append("  v_price numeric(10,2);")
    lines.append("  v_imported int := 0;")
    lines.append("  v_skipped int := 0;")
    lines.append("BEGIN")
    lines.append("  INSERT INTO public.brands (slug, name, color, is_active)")
    lines.append("  VALUES ('sanyo', 'Sanyo', '#1D4ED8', true)")
    lines.append("  ON CONFLICT (slug) DO UPDATE SET is_active = excluded.is_active;")
    lines.append("")
    lines.append("  SELECT id INTO v_type_id FROM public.product_types ORDER BY name LIMIT 1;")
    lines.append("  IF v_type_id IS NULL THEN")
    lines.append("    RAISE EXCEPTION 'Липсва product_types seed.';")
    lines.append("  END IF;")
    lines.append("")
    lines.append("  FOR r IN")
    lines.append("    SELECT *")
    lines.append("    FROM (")
    lines.append("      VALUES")

    value_lines: list[str] = []
    for row in rows:
        value_lines.append(
            "        ("
            + ", ".join([
                sql_lit(row.stock_slug_prefix, "text"),
                sql_lit(row.sheet_name, "text"),
                sql_lit(row.sheet_row, "int"),
                sql_lit(row.brand_db, "text"),
                sql_lit(row.model, "text"),
                sql_lit(row.indoor_serial or None, "text"),
                sql_lit(row.outdoor_serial or None, "text"),
                sql_lit(row.purchase_price, "numeric(10,2)"),
                sql_lit(row.list_price, "numeric(10,2)"),
                sql_lit(build_stock_description(row), "text"),
            ])
            + ")"
        )

    lines.append(",\n".join(value_lines))
    lines.append("    ) AS stage(")
    lines.append("      stock_slug_prefix, sheet_name, sheet_row, brand_name, model,")
    lines.append("      indoor_serial, outdoor_serial, purchase_price, list_price, description")
    lines.append("    )")
    lines.append("    ORDER BY stock_slug_prefix, sheet_name, sheet_row")
    lines.append("  LOOP")
    lines.append("    v_product_id := NULL;")
    lines.append("    v_slug := r.stock_slug_prefix || '-' || r.sheet_name || '-' || r.sheet_row;")
    lines.append("    v_desc := r.description;")
    lines.append("    v_price := coalesce(r.list_price, r.purchase_price, 0);")
    lines.append("")
    lines.append("    IF EXISTS (SELECT 1 FROM public.products WHERE slug = v_slug) THEN")
    lines.append("      v_skipped := v_skipped + 1;")
    lines.append("      CONTINUE;")
    lines.append("    END IF;")
    lines.append("    IF EXISTS (")
    lines.append("      SELECT 1 FROM public.products")
    lines.append("      WHERE description LIKE 'Импорт % VTORA склад, лист ' || r.sheet_name || ' ред ' || r.sheet_row || '%'")
    lines.append("    ) THEN")
    lines.append("      v_skipped := v_skipped + 1;")
    lines.append("      CONTINUE;")
    lines.append("    END IF;")
    lines.append("")
    lines.append("    IF r.indoor_serial IS NOT NULL AND btrim(r.indoor_serial) <> '' THEN")
    lines.append("      SELECT id INTO v_product_id FROM public.products p")
    lines.append("      WHERE upper(btrim(p.indoor_unit_serial)) = upper(btrim(r.indoor_serial))")
    lines.append("      LIMIT 1;")
    lines.append("    END IF;")
    lines.append("    IF v_product_id IS NULL AND r.outdoor_serial IS NOT NULL AND btrim(r.outdoor_serial) <> '' THEN")
    lines.append("      SELECT id INTO v_product_id FROM public.products p")
    lines.append("      WHERE upper(btrim(p.outdoor_unit_serial)) = upper(btrim(r.outdoor_serial))")
    lines.append("      LIMIT 1;")
    lines.append("    END IF;")
    lines.append("    IF v_product_id IS NOT NULL THEN")
    lines.append("      v_skipped := v_skipped + 1;")
    lines.append("      CONTINUE;")
    lines.append("    END IF;")
    lines.append("")
    lines.append("    SELECT id INTO v_brand_id FROM public.brands WHERE name = r.brand_name LIMIT 1;")
    lines.append("    IF v_brand_id IS NULL THEN")
    lines.append("      RAISE WARNING 'Klimatici stock % row %: липсва марка %', r.sheet_name, r.sheet_row, r.brand_name;")
    lines.append("      CONTINUE;")
    lines.append("    END IF;")
    lines.append("")
    lines.append("    v_name := r.brand_name || ' ' || coalesce(nullif(btrim(r.model), ''), 'климатик');")
    lines.append("")
    lines.append("    INSERT INTO public.products (")
    lines.append("      slug, name, brand_id, type_id, model_code, description, price, purchase_price,")
    lines.append("      indoor_unit_serial, outdoor_unit_serial, purchased_at,")
    lines.append("      product_condition, product_region, stock_status, stock_quantity, sold_quantity,")
    lines.append("      is_active, show_in_public_catalog")
    lines.append("    ) VALUES (")
    lines.append("      v_slug, v_name, v_brand_id, v_type_id, nullif(btrim(r.model), ''), v_desc,")
    lines.append("      v_price, r.purchase_price,")
    lines.append("      nullif(btrim(r.indoor_serial), ''), nullif(btrim(r.outdoor_serial), ''), NULL,")
    lines.append("      'used', 'europe', 'in_stock', 1, 0, false, false")
    lines.append("    );")
    lines.append("    v_imported := v_imported + 1;")
    lines.append("  END LOOP;")
    lines.append("")
    lines.append("  RAISE NOTICE 'Klimatici used stock import: imported=%, skipped(existing)=%', v_imported, v_skipped;")
    lines.append("END")
    lines.append("$import$;")
    lines.append("")

    STOCK_OUT_SQL.write_text("\n".join(lines), encoding="utf-8")
    print(f"SQL seed: {STOCK_OUT_SQL}")
    print(f"Rows: {len(rows)}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import Klimatici used AC sales (2023/2022/vtora)")
    parser.add_argument("--year", choices=sorted(IMPORT_CONFIGS), default=None)
    parser.add_argument("--stock", action="store_true", help="Import inventory rows (used, in_stock)")
    parser.add_argument("--preview", action="store_true")
    parser.add_argument("--sql", action="store_true")
    args = parser.parse_args()

    if not args.preview and not args.sql:
        parser.error("Specify --preview and/or --sql")

    if args.stock:
        target_years = [args.year] if args.year else sorted(IMPORT_CONFIGS)
        all_rows: list[ParsedStockRow] = []
        combined_stats: Counter = Counter()
        for year in target_years:
            config = IMPORT_CONFIGS[year]
            if not config.xlsx_path.exists():
                print(f"Missing source file: {config.xlsx_path}", file=sys.stderr)
                return 1
            rows, stats = parse_stock_workbook(config)
            all_rows.extend(rows)
            combined_stats.update(stats)
        if args.preview:
            write_stock_preview(all_rows, combined_stats)
        if args.sql:
            write_stock_sql(all_rows)
        return 0

    year = args.year or "2023"
    config = IMPORT_CONFIGS[year]
    if not config.xlsx_path.exists():
        print(f"Missing source file: {config.xlsx_path}", file=sys.stderr)
        return 1

    rows, stats = parse_workbook(config)

    if args.preview:
        write_preview(rows, stats, config)
    if args.sql:
        write_sql(rows, config)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
