#!/usr/bin/env python3
"""
Импорт на наличност „втора употреба“ (JAPAN) от Klimatici vtora2024.xlsx.

Включва само редове с **черен** шрифт на външния модел (не зелен FF00B050, не червен FF0000),
независимо от фона на клетката.

Използване:
  python backend/scripts/import_klimatici_vtora2024_stock.py --preview
  python backend/scripts/import_klimatici_vtora2024_stock.py --sql

Изход:
  Doc/Klimatici_vtora2024_japan_stock_preview.tsv
  backend/supabase/seeds/0021_klimatici_vtora2024_japan_stock.sql
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from import_klimatici2023_used_sales import (
    ROOT,
    brand_from_sheet,
    cell_str,
    clean_text,
    is_plausible_serial,
    parse_date,
    parse_price,
    sql_lit,
)

XLSX_PATH = ROOT.parent / "Doc" / "Klimatici vtora2024.xlsx"
PREVIEW_TSV = ROOT.parent / "Doc" / "Klimatici_vtora2024_japan_stock_preview.tsv"
WARNINGS_TXT = ROOT.parent / "Doc" / "Klimatici_vtora2024_japan_stock_warnings.txt"
OUT_SQL = ROOT / "backend" / "supabase" / "seeds" / "0021_klimatici_vtora2024_japan_stock.sql"
ROLLBACK_SEED = "0022_rollback_klimatici_vtora2024_japan_stock.sql"
IMPORT_LABEL = "Klimatici vtora2024 JAPAN"
SLUG_PREFIX = "klimatici-vtora2024-japan"

JUNK_OUTDOOR = {"няма табелка", "брак", "brak", "bez tabela", "bez tabela"}


@dataclass(frozen=True)
class SheetLayout:
    outdoor_col: int
    indoor_col: int
    freon_col: int | None = None
    kg_col: int | None = None
    nominal_col: int | None = None
    cooling_kw_col: int | None = None
    heating_kw_col: int | None = None
    heating_minus7_col: int | None = None
    cop_col: int | None = None
    eer_col: int | None = None
    cons_cool_col: int | None = None
    cons_heat_col: int | None = None
    year_col: int | None = None
    notes_col: int | None = None
    extra_v_col: int | None = None
    extra_w_col: int | None = None
    sale_client_col: int | None = None
    sale_date_col: int | None = None
    sale_price_col: int | None = None


SHEET_LAYOUTS: dict[str, SheetLayout] = {
    "Daikin вън": SheetLayout(
        outdoor_col=1,
        indoor_col=2,
        sale_client_col=3,
        sale_date_col=4,
        sale_price_col=5,
        freon_col=5,
        kg_col=6,
        nominal_col=7,
        cooling_kw_col=8,
        heating_kw_col=9,
        heating_minus7_col=10,
        cop_col=11,
        eer_col=12,
        cons_cool_col=13,
        cons_heat_col=14,
        year_col=15,
        notes_col=16,
    ),
    "Mitsubishi": SheetLayout(
        outdoor_col=2,
        indoor_col=3,
        sale_client_col=4,
        sale_price_col=6,
        extra_v_col=4,
        extra_w_col=5,
        freon_col=6,
        kg_col=7,
        nominal_col=8,
    ),
    "Nacional Вън": SheetLayout(
        outdoor_col=2,
        indoor_col=3,
        extra_v_col=4,
        extra_w_col=5,
        freon_col=None,
        kg_col=7,
        nominal_col=8,
        year_col=9,
    ),
    "Panasonic вън": SheetLayout(
        outdoor_col=2,
        indoor_col=3,
        extra_v_col=4,
        extra_w_col=5,
        freon_col=6,
        kg_col=7,
        nominal_col=8,
        year_col=9,
    ),
    "Fujitsu Вън": SheetLayout(
        outdoor_col=2,
        indoor_col=3,
        extra_v_col=4,
        extra_w_col=5,
        freon_col=6,
        kg_col=7,
    ),
}

DEFAULT_LAYOUT = SheetLayout(
    outdoor_col=2,
    indoor_col=3,
    extra_v_col=4,
    extra_w_col=5,
    freon_col=6,
    kg_col=7,
    nominal_col=8,
    cooling_kw_col=9,
    heating_kw_col=10,
    heating_minus7_col=11,
    cop_col=12,
    eer_col=13,
    cons_cool_col=14,
    cons_heat_col=15,
    year_col=16,
    notes_col=17,
)


@dataclass
class ParsedStockRow:
    sheet_name: str
    sheet_row: int
    brand_db: str
    outdoor_model: str
    indoor_serial: str
    purchase_price: float | None
    list_price: float | None
    refrigerant: str
    weight_kg: float | None
    nominal_label: str
    cooling_kw: float | None
    heating_kw: float | None
    heating_minus7_kw: float | None
    cop: float | None
    eer: float | None
    cons_cool_kw: float | None
    cons_heat_kw: float | None
    manufacture_year: str
    notes: str
    sale_date: str
    extra_v: str
    extra_w: str
    btu: int | None
    description: str
    warnings: list[str] = field(default_factory=list)


def looks_like_refrigerant(val) -> bool:
    if val is None or val == "":
        return False
    s = clean_text(str(val)).upper().replace(" ", "")
    return bool(re.fullmatch(r"R\d{2,3}A?", s))


def looks_like_client_text(text: str) -> bool:
    t = clean_text(text)
    if not t:
        return False
    if parse_date(t):
        return False
    if parse_price(t) is not None and not re.search(r"[A-Za-zА-Яа-я]", t):
        return False
    return bool(re.search(r"\d{9,}", t) or re.search(r"[A-Za-zА-Яа-я]{3,}", t))


def read_refrigerant(ws, row: int, col: int | None, reserved_cols: set[int]) -> str:
    if col is None or col in reserved_cols:
        return ""
    val = ws.cell(row, col).value
    if looks_like_refrigerant(val):
        return clean_text(str(val)).upper()
    return ""


def read_spec_decimal(ws, row: int, col: int | None, reserved_cols: set[int]) -> float | None:
    if col is None or col in reserved_cols:
        return None
    val = ws.cell(row, col).value
    if val is None or val == "":
        return None
    if looks_like_refrigerant(val):
        return None
    price = parse_price(val)
    if price is not None and price >= 100 and not parse_decimal(val):
        return None
    return parse_decimal(val)


def extract_sale_fields(ws, row: int, layout: SheetLayout) -> tuple[str, str, float | None, set[int]]:
    reserved: set[int] = set()
    client = date_note = ""
    list_price = None

    if layout.sale_client_col:
        raw = cell_str(ws.cell(row, layout.sale_client_col).value)
        if looks_like_client_text(raw):
            client = clean_text(raw)
            reserved.add(layout.sale_client_col)

    if layout.sale_date_col:
        raw = cell_str(ws.cell(row, layout.sale_date_col).value)
        parsed_date = parse_date(raw)
        if parsed_date:
            date_note = parsed_date
            reserved.add(layout.sale_date_col)
        elif looks_like_client_text(raw):
            if client:
                client = f"{client} · {clean_text(raw)}"
            else:
                client = clean_text(raw)
            reserved.add(layout.sale_date_col)

    if layout.sale_price_col:
        raw = ws.cell(row, layout.sale_price_col).value
        price = parse_price(raw)
        if price is not None and price >= 100:
            list_price = price
            reserved.add(layout.sale_price_col)

    return client, date_note, list_price, reserved


def font_color_kind(cell) -> str:
    """black = наличност; green/red = продадени/резервирани."""
    fc = cell.font.color
    if fc is None:
        return "black"
    if fc.type == "rgb" and fc.rgb:
        rgb = str(fc.rgb).upper()
        if "00B050" in rgb:
            return "green"
        if rgb in ("FFFF0000", "FF0000") or rgb.endswith("FF0000"):
            return "red"
        return "black"
    if fc.type == "theme":
        return "black" if fc.theme in (0, 1) else f"theme:{fc.theme}"
    return "black"


def is_stock_row(outdoor_cell) -> bool:
    return font_color_kind(outdoor_cell) == "black"


def parse_decimal(val) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        n = round(float(val), 2)
        return n if n > 0 else None
    s = str(val).strip().replace(",", ".")
    try:
        n = round(float(s), 2)
        return n if n > 0 else None
    except ValueError:
        return None


def parse_int(val) -> int | None:
    if val is None or val == "":
        return None
    if isinstance(val, int):
        return val if val > 0 else None
    if isinstance(val, float) and val.is_integer():
        v = int(val)
        return v if v > 0 else None
    s = re.sub(r"[^\d]", "", str(val).strip())
    if not s:
        return None
    v = int(s)
    return v if v > 0 else None


def infer_btu(outdoor_model: str, cooling_kw: float | None, nominal_label: str) -> int | None:
    m = re.match(r"^(\d{2,3})", outdoor_model.strip())
    if m:
        prefix = int(m.group(1))
        mapping = {
            18: 7,
            20: 7,
            22: 9,
            24: 9,
            25: 9,
            28: 10,
            35: 12,
            36: 12,
            40: 14,
            42: 15,
            50: 18,
            56: 18,
            60: 24,
            70: 24,
        }
        for key in sorted(mapping, reverse=True):
            if prefix >= key:
                return mapping[key] * 1000
    btu_from_nom = parse_int(nominal_label)
    if btu_from_nom and btu_from_nom >= 7:
        return btu_from_nom * 1000 if btu_from_nom < 100 else btu_from_nom
    if cooling_kw:
        if cooling_kw <= 2.6:
            return 9000
        if cooling_kw <= 3.7:
            return 12000
        if cooling_kw <= 5.3:
            return 18000
        if cooling_kw <= 6.5:
            return 21000
        return 24000
    return None


def cell_text(ws, row: int, col: int | None) -> str:
    if col is None:
        return ""
    return clean_text(cell_str(ws.cell(row, col).value))


def build_description(row: ParsedStockRow) -> str:
    parts = [f"Импорт {IMPORT_LABEL} склад, лист {row.sheet_name} ред {row.sheet_row}"]
    spec_bits: list[str] = []
    if row.refrigerant:
        spec_bits.append(f"фреон: {row.refrigerant}")
    if row.weight_kg is not None:
        spec_bits.append(f"тегло: {row.weight_kg} kg")
    if row.cooling_kw is not None:
        spec_bits.append(f"охлаждане: {row.cooling_kw} kW")
    if row.heating_kw is not None:
        spec_bits.append(f"отопление: {row.heating_kw} kW")
    if row.heating_minus7_kw is not None:
        spec_bits.append(f"отопление -7°C: {row.heating_minus7_kw} kW")
    if row.cop is not None:
        spec_bits.append(f"COP: {row.cop}")
    if row.eer is not None:
        spec_bits.append(f"EER: {row.eer}")
    if row.cons_cool_kw is not None:
        spec_bits.append(f"консум. охл.: {row.cons_cool_kw} kW")
    if row.cons_heat_kw is not None:
        spec_bits.append(f"консум. отопл.: {row.cons_heat_kw} kW")
    if row.manufacture_year:
        spec_bits.append(f"година: {row.manufacture_year}")
    if row.sale_date:
        spec_bits.append(f"дата: {row.sale_date}")
    if row.nominal_label:
        spec_bits.append(f"номинал: {row.nominal_label}")
    if row.extra_v:
        spec_bits.append(f"v: {row.extra_v}")
    if row.extra_w:
        spec_bits.append(f"w: {row.extra_w}")
    if spec_bits:
        parts.append(" · ".join(spec_bits))
    if row.notes:
        parts.append(row.notes)
    return " · ".join(parts)


def parse_workbook(xlsx_path: Path) -> tuple[list[ParsedStockRow], Counter]:
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    wb_fmt = openpyxl.load_workbook(str(xlsx_path), data_only=False)
    rows: list[ParsedStockRow] = []
    stats: Counter = Counter()

    for sheet_name in wb.sheetnames:
        brand_db = brand_from_sheet(sheet_name)
        if not brand_db:
            stats[f"unknown_sheet:{sheet_name}"] += 1
            continue

        layout = SHEET_LAYOUTS.get(sheet_name, DEFAULT_LAYOUT)
        ws = wb[sheet_name]
        ws_fmt = wb_fmt[sheet_name]

        for r in range(2, ws.max_row + 1):
            outdoor = cell_str(ws.cell(r, layout.outdoor_col).value)
            indoor = cell_str(ws.cell(r, layout.indoor_col).value)
            if not outdoor and not indoor:
                continue

            outdoor_cell = ws_fmt.cell(r, layout.outdoor_col)
            if not is_stock_row(outdoor_cell):
                stats["skipped_colored_font"] += 1
                continue

            if outdoor.lower() in JUNK_OUTDOOR:
                stats["skipped_junk"] += 1
                continue
            if not is_plausible_serial(outdoor) and not is_plausible_serial(indoor):
                stats["skipped_implausible_serial"] += 1
                continue

            sale_client, sale_date, sale_list_price, reserved_cols = extract_sale_fields(ws, r, layout)

            purchase_price = None
            list_price = sale_list_price
            if layout.extra_v_col and layout.extra_v_col not in reserved_cols:
                purchase_price = parse_price(ws.cell(r, layout.extra_v_col).value)
            if layout.extra_w_col and layout.extra_w_col not in reserved_cols and list_price is None:
                list_price = parse_price(ws.cell(r, layout.extra_w_col).value)

            refrigerant = read_refrigerant(ws, r, layout.freon_col, reserved_cols)
            weight_kg = read_spec_decimal(ws, r, layout.kg_col, reserved_cols)
            nominal_label = cell_text(ws, r, layout.nominal_col) if layout.nominal_col not in reserved_cols else ""
            cooling_kw = read_spec_decimal(ws, r, layout.cooling_kw_col, reserved_cols)
            heating_kw = read_spec_decimal(ws, r, layout.heating_kw_col, reserved_cols)
            heating_minus7_kw = read_spec_decimal(ws, r, layout.heating_minus7_col, reserved_cols)
            cop = read_spec_decimal(ws, r, layout.cop_col, reserved_cols)
            eer = read_spec_decimal(ws, r, layout.eer_col, reserved_cols)
            cons_cool_kw = read_spec_decimal(ws, r, layout.cons_cool_col, reserved_cols)
            cons_heat_kw = read_spec_decimal(ws, r, layout.cons_heat_col, reserved_cols)
            manufacture_year = cell_text(ws, r, layout.year_col) if layout.year_col not in reserved_cols else ""
            notes = cell_text(ws, r, layout.notes_col) if layout.notes_col not in reserved_cols else ""
            extra_v = cell_text(ws, r, layout.extra_v_col) if layout.extra_v_col and layout.extra_v_col not in reserved_cols else ""
            extra_w = cell_text(ws, r, layout.extra_w_col) if layout.extra_w_col and layout.extra_w_col not in reserved_cols else ""

            note_parts = [p for p in (sale_client, sale_date, notes) if p]
            notes = " · ".join(note_parts)

            btu = infer_btu(outdoor, cooling_kw, nominal_label)
            warnings: list[str] = []
            if list_price is None and purchase_price is None:
                warnings.append("missing_prices")

            parsed = ParsedStockRow(
                sheet_name=sheet_name,
                sheet_row=r,
                brand_db=brand_db,
                outdoor_model=outdoor,
                indoor_serial=indoor,
                purchase_price=purchase_price,
                list_price=list_price,
                refrigerant=refrigerant,
                weight_kg=weight_kg,
                nominal_label=nominal_label,
                cooling_kw=cooling_kw,
                heating_kw=heating_kw,
                heating_minus7_kw=heating_minus7_kw,
                cop=cop,
                eer=eer,
                cons_cool_kw=cons_cool_kw,
                cons_heat_kw=cons_heat_kw,
                manufacture_year=manufacture_year,
                notes=notes,
                sale_date=sale_date,
                extra_v=extra_v,
                extra_w=extra_w,
                btu=btu,
                description="",
                warnings=warnings,
            )
            parsed.description = build_description(parsed)
            rows.append(parsed)
            stats[f"sheet:{sheet_name}"] += 1

    return rows, stats


def write_preview(rows: list[ParsedStockRow], stats: Counter) -> None:
    PREVIEW_TSV.parent.mkdir(parents=True, exist_ok=True)
    with PREVIEW_TSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow([
            "sheet", "row", "brand", "outdoor", "indoor", "purchase", "list",
            "freon", "kg", "cool_kw", "heat_kw", "cop", "eer", "year", "btu", "notes", "description",
        ])
        for row in rows:
            w.writerow([
                row.sheet_name, row.sheet_row, row.brand_db, row.outdoor_model, row.indoor_serial,
                row.purchase_price if row.purchase_price is not None else "",
                row.list_price if row.list_price is not None else "",
                row.refrigerant, row.weight_kg if row.weight_kg is not None else "",
                row.cooling_kw if row.cooling_kw is not None else "",
                row.heating_kw if row.heating_kw is not None else "",
                row.cop if row.cop is not None else "",
                row.eer if row.eer is not None else "",
                row.manufacture_year, row.btu if row.btu is not None else "",
                row.notes, row.description,
            ])

    warn_counter = Counter()
    for row in rows:
        for warning in row.warnings:
            warn_counter[warning] += 1

    lines = [
        f"Source: {XLSX_PATH}",
        f"Black-font stock rows: {len(rows)}",
        "",
        "Per sheet:",
        *[f"  {k}: {v}" for k, v in sorted(stats.items())],
        "",
        "Warnings:",
        *[f"  {k}: {v}" for k, v in sorted(warn_counter.items())],
    ]
    WARNINGS_TXT.write_text("\n".join(lines), encoding="utf-8")
    print(f"Preview: {PREVIEW_TSV}")
    print(f"Warnings: {WARNINGS_TXT}")
    print(f"Rows: {len(rows)}")


def write_sql(rows: list[ParsedStockRow]) -> None:
    OUT_SQL.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = [
        "-- =====================================================================",
        "-- Seed: Klimatici vtora2024 — JAPAN наличност (черен шрифт в Excel)",
        "-- =====================================================================",
        f"-- Редове: {len(rows)}",
        "-- product_condition = used, product_region = japan, stock_status = in_stock",
        f"-- Идемпотентност: slug + description LIKE 'Импорт {IMPORT_LABEL} склад, лист % ред %'",
        f"-- Rollback: seeds/{ROLLBACK_SEED}",
        "-- ВАЖНО: Един DO блок — пусни целия файл (Ctrl+A → Run).",
        "-- =====================================================================",
        "",
        "DO $import$",
        "DECLARE",
        "  r RECORD;",
        "  v_brand_id uuid;",
        "  v_type_id uuid;",
        "  v_product_id uuid;",
        "  v_slug text;",
        "  v_name text;",
        "  v_desc text;",
        "  v_price numeric(10,2);",
        "  v_imported int := 0;",
        "  v_skipped int := 0;",
        "BEGIN",
        "  SELECT id INTO v_type_id FROM public.product_types ORDER BY name LIMIT 1;",
        "  IF v_type_id IS NULL THEN",
        "    RAISE EXCEPTION 'Липсва product_types seed.';",
        "  END IF;",
        "",
        "  FOR r IN",
        "    SELECT *",
        "    FROM (",
        "      VALUES",
    ]

    value_lines: list[str] = []
    for row in rows:
        value_lines.append(
            "        ("
            + ", ".join([
                sql_lit(row.sheet_name, "text"),
                sql_lit(row.sheet_row, "int"),
                sql_lit(row.brand_db, "text"),
                sql_lit(row.outdoor_model, "text"),
                sql_lit(row.indoor_serial or None, "text"),
                sql_lit(row.purchase_price, "numeric(10,2)"),
                sql_lit(row.list_price, "numeric(10,2)"),
                sql_lit(row.sale_date or None, "date"),
                sql_lit(row.description, "text"),
                sql_lit(row.refrigerant or None, "text"),
                sql_lit(row.weight_kg, "numeric(7,2)"),
                sql_lit(row.cooling_kw, "numeric(6,2)"),
                sql_lit(row.heating_kw, "numeric(6,2)"),
                sql_lit(row.cop, "numeric(6,2)"),
                sql_lit(row.eer, "numeric(6,2)"),
                sql_lit(row.btu, "int"),
            ])
            + ")"
        )

    lines.append(",\n".join(value_lines))
    lines.extend([
        "    ) AS stage(",
        "      sheet_name, sheet_row, brand_name, outdoor_model, indoor_serial,",
        "      purchase_price, list_price, purchased_at, description,",
        "      refrigerant, weight_kg, cooling_kw, heating_kw, cop, eer, btu",
        "    )",
        "    ORDER BY sheet_name, sheet_row",
        "  LOOP",
        "    v_product_id := NULL;",
        f"    v_slug := '{SLUG_PREFIX}-' || r.sheet_name || '-' || r.sheet_row;",
        "    v_desc := r.description;",
        "    v_price := coalesce(r.list_price, r.purchase_price, 0);",
        "",
        "    IF EXISTS (SELECT 1 FROM public.products WHERE slug = v_slug) THEN",
        "      v_skipped := v_skipped + 1;",
        "      CONTINUE;",
        "    END IF;",
        f"    IF EXISTS (SELECT 1 FROM public.products WHERE description LIKE 'Импорт {IMPORT_LABEL} склад, лист ' || r.sheet_name || ' ред ' || r.sheet_row || '%') THEN",
        "      v_skipped := v_skipped + 1;",
        "      CONTINUE;",
        "    END IF;",
        "",
        "    IF r.indoor_serial IS NOT NULL AND btrim(r.indoor_serial) <> '' THEN",
        "      SELECT id INTO v_product_id FROM public.products p",
        "      WHERE upper(btrim(p.indoor_unit_serial)) = upper(btrim(r.indoor_serial))",
        "      LIMIT 1;",
        "    END IF;",
        "    IF v_product_id IS NULL AND r.outdoor_model IS NOT NULL AND btrim(r.outdoor_model) <> '' THEN",
        "      SELECT id INTO v_product_id FROM public.products p",
        "      WHERE upper(btrim(p.outdoor_unit_serial)) = upper(btrim(r.outdoor_model))",
        "         OR upper(btrim(p.model_code)) = upper(btrim(r.outdoor_model))",
        "      LIMIT 1;",
        "    END IF;",
        "    IF v_product_id IS NOT NULL THEN",
        "      v_skipped := v_skipped + 1;",
        "      CONTINUE;",
        "    END IF;",
        "",
        "    SELECT id INTO v_brand_id FROM public.brands WHERE name = r.brand_name LIMIT 1;",
        "    IF v_brand_id IS NULL THEN",
        "      RAISE WARNING 'Klimatici vtora2024 % row %: липсва марка %', r.sheet_name, r.sheet_row, r.brand_name;",
        "      CONTINUE;",
        "    END IF;",
        "",
        "    v_name := r.brand_name || ' ' || coalesce(nullif(btrim(r.outdoor_model), ''), 'климатик');",
        "",
        "    INSERT INTO public.products (",
        "      slug, name, brand_id, type_id, model_code, description, price, purchase_price,",
        "      indoor_unit_serial, outdoor_unit_serial, purchased_at,",
        "      product_condition, product_region, stock_status, stock_quantity, sold_quantity,",
        "      is_active, show_in_public_catalog",
        "    ) VALUES (",
        "      v_slug, v_name, v_brand_id, v_type_id, nullif(btrim(r.outdoor_model), ''), v_desc,",
        "      v_price, r.purchase_price,",
        "      nullif(btrim(r.indoor_serial), ''), nullif(btrim(r.outdoor_model), ''), r.purchased_at,",
        "      'used', 'japan', 'in_stock', 1, 0, false, false",
        "    ) RETURNING id INTO v_product_id;",
        "",
        "    INSERT INTO public.product_specs (",
        "      product_id, refrigerant, cooling_power_kw, heating_power_kw,",
        "      seer, scop, btu, weight_outdoor_kg",
        "    ) VALUES (",
        "      v_product_id, nullif(btrim(r.refrigerant), ''), r.cooling_kw, r.heating_kw,",
        "      r.eer, r.cop, r.btu, r.weight_kg",
        "    )",
        "    ON CONFLICT (product_id) DO UPDATE SET",
        "      refrigerant = excluded.refrigerant,",
        "      cooling_power_kw = excluded.cooling_power_kw,",
        "      heating_power_kw = excluded.heating_power_kw,",
        "      seer = excluded.seer,",
        "      scop = excluded.scop,",
        "      btu = excluded.btu,",
        "      weight_outdoor_kg = excluded.weight_outdoor_kg;",
        "",
        "    v_imported := v_imported + 1;",
        "  END LOOP;",
        "",
        "  RAISE NOTICE 'Klimatici vtora2024 JAPAN stock: imported=%, skipped(existing)=%', v_imported, v_skipped;",
        "END",
        "$import$;",
        "",
    ])

    OUT_SQL.write_text("\n".join(lines), encoding="utf-8")
    print(f"SQL seed: {OUT_SQL}")
    print(f"Rows: {len(rows)}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import Klimatici vtora2024 JAPAN stock (black font rows)")
    parser.add_argument("--preview", action="store_true")
    parser.add_argument("--sql", action="store_true")
    args = parser.parse_args()

    if not args.preview and not args.sql:
        parser.error("Specify --preview and/or --sql")

    if not XLSX_PATH.exists():
        print(f"Missing source file: {XLSX_PATH}", file=sys.stderr)
        return 1

    rows, stats = parse_workbook(XLSX_PATH)

    if args.preview:
        write_preview(rows, stats)
    if args.sql:
        write_sql(rows)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
